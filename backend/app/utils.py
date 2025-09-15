import os
import zipfile
import asyncio
from typing import List, Dict, Any
from fpdf import FPDF
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np
from datetime import datetime
import tempfile
import requests
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import ResultadoCorrecao, PDFInfo, ConfiguracaoSistema

class GeradorPDF:
    """Classe para geração de PDFs dos boletins"""
    
    def __init__(self):
        self.config = ConfiguracaoSistema()
        self.temp_dir = tempfile.mkdtemp()
        
        # URLs das logos (do GitHub)
        self.logo_acafe_url = "https://raw.githubusercontent.com/JulioFloripa/CorretorACAFE/main/logo-acafe.png"
        self.logo_fleming_url = "https://raw.githubusercontent.com/JulioFloripa/CorretorACAFE/main/logo_fleming.png"
    
    async def gerar_todos_pdfs_async(self, resultado_processamento: Dict[str, Any] ) -> List[PDFInfo]:
        """Gera PDFs para todos os alunos de forma assíncrona"""
        
        resultados = resultado_processamento["resultados"]
        estatisticas = resultado_processamento["estatisticas"]
        ranking = resultado_processamento["ranking"]
        
        # Gerar PDFs em lotes
        lote_size = 10
        pdfs_info = []
        
        for i in range(0, len(resultados), lote_size):
            lote = resultados[i:i + lote_size]
            lote_pdfs = await asyncio.gather(*[
                self._gerar_pdf_individual_async(resultado, estatisticas, ranking)
                for resultado in lote
            ])
            pdfs_info.extend(lote_pdfs)
        
        return pdfs_info
    
    async def _gerar_pdf_individual_async(self, resultado: ResultadoCorrecao, estatisticas: Any, ranking: List[Any]) -> PDFInfo:
        """Gera PDF individual para um aluno"""
        
        # Encontrar posição no ranking
        posicao = next((i + 1 for i, r in enumerate(ranking) if r.id == resultado.aluno.id), 0)
        
        # Nome do arquivo
        nome_arquivo = f"Boletim_{resultado.aluno.nome.replace(' ', '_')}.pdf"
        caminho_pdf = os.path.join(self.temp_dir, nome_arquivo)
        
        # Gerar PDF
        await self._criar_pdf_boletim(resultado, estatisticas, posicao, caminho_pdf)
        
        # Obter tamanho do arquivo
        tamanho_bytes = os.path.getsize(caminho_pdf)
        
        return PDFInfo(
            aluno_id=resultado.aluno.id,
            nome_aluno=resultado.aluno.nome,
            nome_arquivo=nome_arquivo,
            caminho=caminho_pdf,
            tamanho_bytes=tamanho_bytes
        )
    
    async def _criar_pdf_boletim(self, resultado: ResultadoCorrecao, estatisticas: Any, posicao: int, caminho_pdf: str):
        """Cria o PDF do boletim individual"""
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        
        # Cabeçalho com logos
        await self._adicionar_cabecalho(pdf)
        
        # Informações do aluno
        self._adicionar_info_aluno(pdf, resultado, posicao)
        
        # Resumo de performance
        self._adicionar_resumo_performance(pdf, resultado, estatisticas)
        
        # Desempenho por disciplina
        self._adicionar_desempenho_disciplinas(pdf, resultado)
        
        # Gráfico de performance (se possível)
        grafico_path = await self._gerar_grafico_performance(resultado)
        if grafico_path:
            self._adicionar_grafico(pdf, grafico_path)
        
        # Rodapé
        self._adicionar_rodape(pdf)
        
        # Salvar PDF
        pdf.output(caminho_pdf)
    
    async def _adicionar_cabecalho(self, pdf: FPDF):
        """Adiciona cabeçalho com logos"""
        
        # Fundo verde ACAFE
        pdf.set_fill_color(46, 125, 50)  # Verde ACAFE
        pdf.rect(10, 10, 190, 30, 'F')
        
        # Título centralizado
        pdf.set_text_color(255, 255, 255)  # Branco
        pdf.set_font('Arial', 'B', 18)
        pdf.set_xy(10, 20)
        pdf.cell(190, 10, 'SIMULADO ACAFE - COLEGIO FLEMING', 0, 1, 'C')
        
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', '', 12)
        pdf.set_xy(10, 30)
        pdf.cell(190, 8, 'Sistema Inteligente de Correcao de Simulados', 0, 1, 'C')
        
        # Reset cor do texto
        pdf.set_text_color(0, 0, 0)
        pdf.ln(15)
    
    def _adicionar_info_aluno(self, pdf: FPDF, resultado: ResultadoCorrecao, posicao: int):
        """Adiciona informações básicas do aluno"""
        
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'INFORMACOES DO ALUNO', 0, 1, 'L')
        pdf.ln(5)
        
        # Caixa com informações
        pdf.set_fill_color(240, 248, 255)  # Azul claro
        pdf.rect(10, pdf.get_y(), 190, 25, 'F')
        
        pdf.set_font('Arial', '', 12)
        y_start = pdf.get_y() + 5
        
        pdf.set_xy(15, y_start)
        pdf.cell(0, 6, f'Nome: {resultado.aluno.nome}', 0, 1, 'L')
        
        pdf.set_xy(15, y_start + 8)
        pdf.cell(90, 6, f'ID: {resultado.aluno.id}', 0, 0, 'L')
        pdf.cell(90, 6, f'Posicao no Ranking: {posicao}', 0, 1, 'L')
        
        if resultado.aluno.sede:
            pdf.set_xy(15, y_start + 16)
            pdf.cell(0, 6, f'Sede: {resultado.aluno.sede}', 0, 1, 'L')
        
        pdf.ln(15)
    
    def _adicionar_resumo_performance(self, pdf: FPDF, resultado: ResultadoCorrecao, estatisticas: Any):
        """Adiciona resumo de performance"""
        
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'RESUMO DE PERFORMANCE', 0, 1, 'L')
        pdf.ln(5)
        
        # Determinar cor baseada na performance
        nota = resultado.nota_percentual
        if nota >= 85:
            cor_fundo = (76, 175, 80)  # Verde
            status = "EXCELENTE"
        elif nota >= 70:
            cor_fundo = (156, 204, 101)  # Verde claro
            status = "BOM"
        elif nota >= 50:
            cor_fundo = (255, 193, 7)  # Amarelo
            status = "REGULAR"
        else:
            cor_fundo = (244, 67, 54)  # Vermelho
            status = "PRECISA MELHORAR"
        
        # Caixa colorida com nota
        pdf.set_fill_color(*cor_fundo)
        pdf.rect(10, pdf.get_y(), 190, 20, 'F')
        
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', 'B', 16)
        pdf.set_xy(15, pdf.get_y() + 6)
        pdf.cell(0, 8, f'NOTA: {nota:.1f}% - {status}', 0, 1, 'C')
        
        pdf.set_text_color(0, 0, 0)
        pdf.ln(10)
        
        # Detalhes
        pdf.set_font('Arial', '', 11)
        total_questoes = resultado.acertos + resultado.erros
        
        detalhes = [
            f'Acertos: {resultado.acertos}/{total_questoes}',
            f'Percentual de Acerto: {nota:.1f}%',
            f'Media da Turma: {estatisticas.gerais.media_geral:.1f}%',
            f'Diferenca da Media: {nota - estatisticas.gerais.media_geral:+.1f}%'
        ]
        
        for detalhe in detalhes:
            pdf.cell(0, 6, detalhe, 0, 1, 'L')
        
        pdf.ln(10)
    
    def _adicionar_desempenho_disciplinas(self, pdf: FPDF, resultado: ResultadoCorrecao):
        """Adiciona tabela de desempenho por disciplina"""
        
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'DESEMPENHO POR DISCIPLINA', 0, 1, 'L')
        pdf.ln(5)
        
        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(46, 125, 50)  # Verde ACAFE
        pdf.set_text_color(255, 255, 255)
        
        col_widths = [80, 30, 30, 50]
        headers = ['Disciplina', 'Acertos', 'Total', 'Percentual']
        
        x_start = 10
        for i, header in enumerate(headers):
            pdf.set_xy(x_start + sum(col_widths[:i]), pdf.get_y())
            pdf.cell(col_widths[i], 8, header, 1, 0, 'C', True)
        
        pdf.ln(8)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 9)
        
        # Dados das disciplinas
        for disciplina, stats in resultado.desempenho_por_disciplina.items():
            # Cor alternada para linhas
            if len(resultado.desempenho_por_disciplina) % 2 == 0:
                pdf.set_fill_color(248, 249, 250)
            else:
                pdf.set_fill_color(255, 255, 255)
            
            percentual = stats['percentual']
            
            # Cor baseada na performance
            if percentual >= 70:
                pdf.set_text_color(46, 125, 50)  # Verde
            elif percentual >= 50:
                pdf.set_text_color(255, 152, 0)  # Laranja
            else:
                pdf.set_text_color(244, 67, 54)  # Vermelho
            
            dados = [
                disciplina[:25],  # Limitar nome da disciplina
                str(stats['acertos']),
                str(stats['total']),
                f"{percentual:.1f}%"
            ]
            
            for i, dado in enumerate(dados):
                pdf.set_xy(x_start + sum(col_widths[:i]), pdf.get_y())
                pdf.cell(col_widths[i], 6, dado, 1, 0, 'C', True)
            
            pdf.ln(6)
        
        pdf.set_text_color(0, 0, 0)
        pdf.ln(10)
    
    async def _gerar_grafico_performance(self, resultado: ResultadoCorrecao) -> str:
        """Gera gráfico de performance por disciplina"""
        
        try:
            # Configurar estilo
            plt.style.use('default')
            fig, ax = plt.subplots(figsize=(8, 5))
            
            # Dados para o gráfico
            disciplinas = []
            percentuais = []
            cores = []
            
            for disciplina, stats in resultado.desempenho_por_disciplina.items():
                disciplinas.append(disciplina[:15])  # Limitar nome
                percentual = stats['percentual']
                percentuais.append(percentual)
                
                # Cor baseada na performance
                if percentual >= 70:
                    cores.append('#2E7D32')  # Verde ACAFE
                elif percentual >= 50:
                    cores.append('#FF9800')  # Laranja
                else:
                    cores.append('#F44336')  # Vermelho
            
            # Criar gráfico de barras
            bars = ax.bar(disciplinas, percentuais, color=cores, alpha=0.8)
            
            # Personalizar gráfico
            ax.set_ylabel('Percentual de Acertos (%)', fontsize=12)
            ax.set_title(f'Desempenho por Disciplina - {resultado.aluno.nome}', fontsize=14, fontweight='bold')
            ax.set_ylim(0, 100)
            
            # Adicionar valores nas barras
            for bar, percentual in zip(bars, percentuais):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                       f'{percentual:.1f}%', ha='center', va='bottom', fontsize=10)
            
            # Linha da média (70%)
            ax.axhline(y=70, color='red', linestyle='--', alpha=0.7, label='Meta (70%)')
            ax.legend()
            
            # Rotacionar labels do eixo x
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            
            # Salvar gráfico
            grafico_path = os.path.join(self.temp_dir, f'grafico_{resultado.aluno.id}.png')
            plt.savefig(grafico_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return grafico_path
            
        except Exception as e:
            print(f"Erro ao gerar gráfico: {e}")
            return None
    
    def _adicionar_grafico(self, pdf: FPDF, grafico_path: str):
        """Adiciona gráfico ao PDF"""
        
        try:
            # Verificar se há espaço suficiente
            if pdf.get_y() > 200:
                pdf.add_page()
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, 'GRAFICO DE DESEMPENHO', 0, 1, 'L')
            pdf.ln(5)
            
            # Adicionar imagem
            pdf.image(grafico_path, x=10, y=pdf.get_y(), w=190)
            pdf.ln(100)  # Espaço após o gráfico
            
        except Exception as e:
            print(f"Erro ao adicionar gráfico ao PDF: {e}")
    
    def _adicionar_rodape(self, pdf: FPDF):
        """Adiciona rodapé ao PDF"""
        
        # Posicionar no final da página
        pdf.set_y(-30)
        
        # Linha separadora
        pdf.set_draw_color(46, 125, 50)  # Verde ACAFE
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)
        
        # Texto do rodapé
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(128, 128, 128)  # Cinza
        
        data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")
        rodape_texto = f'Boletim gerado em {data_geracao} | Corretor ACAFE Fleming v2.0 | Logos Oficiais'
        
        pdf.cell(0, 5, rodape_texto, 0, 1, 'C')
    
    async def criar_zip_pdfs(self, pdfs_info: List[PDFInfo]) -> str:
        """Cria arquivo ZIP com todos os PDFs"""
        
        zip_path = os.path.join(self.temp_dir, 'boletins_simulado.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for pdf_info in pdfs_info:
                if os.path.exists(pdf_info.caminho):
                    zipf.write(pdf_info.caminho, pdf_info.nome_arquivo)
        
        return zip_path
    
    def criar_template_excel(self) -> str:
        """Cria template Excel para download"""
        
        template_path = os.path.join(self.temp_dir, 'template_simulado_acafe.xlsx')
        
        # Criar workbook
        wb = openpyxl.Workbook()
        
        # Remover sheet padrão
        wb.remove(wb.active)
        
        # Criar aba RESPOSTAS
        ws_respostas = wb.create_sheet("RESPOSTAS")
        self._criar_aba_respostas(ws_respostas)
        
        # Criar aba GABARITO
        ws_gabarito = wb.create_sheet("GABARITO")
        self._criar_aba_gabarito(ws_gabarito)
        
        # Criar aba INSTRUÇÕES
        ws_instrucoes = wb.create_sheet("INSTRUÇÕES")
        self._criar_aba_instrucoes(ws_instrucoes)
        
        # Salvar
        wb.save(template_path)
        
        return template_path
    
    def _criar_aba_respostas(self, ws):
        """Cria aba RESPOSTAS do template"""
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'Sede', 'Idioma escolhido']
        
        # Adicionar colunas de questões (1-70)
        for i in range(1, 71):
            headers.append(f'Questão {i:02d}')
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Exemplos de dados
        exemplos = [
            ['2024001', 'João Silva Santos', 'Unidade Centro', 'Inglês'] + ['A'] * 70,
            ['2024002', 'Maria Oliveira Costa', 'Unidade Norte', 'Espanhol'] + ['B'] * 70,
            ['2024003', 'Pedro Souza Lima', 'Unidade Sul', 'Inglês'] + ['C'] * 70
        ]
        
        for row, exemplo in enumerate(exemplos, 2):
            for col, valor in enumerate(exemplo, 1):
                ws.cell(row=row, column=col, value=valor)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 25  # Nome
        ws.column_dimensions['C'].width = 15  # Sede
        ws.column_dimensions['D'].width = 15  # Idioma
        
        # Questões com largura menor
        for col in range(5, 75):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 8
    
    def _criar_aba_gabarito(self, ws):
        """Cria aba GABARITO do template"""
        
        # Cabeçalhos
        headers = ['Questão', 'Disciplina', 'Resposta']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Exemplo de gabarito
        disciplinas = [
            'Biologia', 'Química', 'Física', 'Matemática',
            'História', 'Geografia', 'Língua Portuguesa e Literatura'
        ]
        
        respostas_exemplo = ['A', 'B', 'C', 'D', 'E']
        
        row = 2
        for questao in range(1, 71):
            if questao <= 56:
                # Questões regulares
                disciplina = disciplinas[(questao - 1) % len(disciplinas)]
            else:
                # Questões de línguas (57-63)
                if questao <= 63:
                    disciplina = 'Inglês'
                else:
                    disciplina = 'Espanhol'
            
            resposta = respostas_exemplo[(questao - 1) % len(respostas_exemplo)]
            
            ws.cell(row=row, column=1, value=questao)
            ws.cell(row=row, column=2, value=disciplina)
            ws.cell(row=row, column=3, value=resposta)
            row += 1
        
        # Adicionar questões de Espanhol (57-63)
        for questao in range(57, 64):
            ws.cell(row=row, column=1, value=questao)
            ws.cell(row=row, column=2, value='Espanhol')
            ws.cell(row=row, column=3, value=respostas_exemplo[(questao - 57) % len(respostas_exemplo)])
            row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
    
    def _criar_aba_instrucoes(self, ws):
        """Cria aba INSTRUÇÕES do template"""
        
        instrucoes = [
            "INSTRUÇÕES PARA USO DO TEMPLATE SIMULADO ACAFE",
            "",
            "1. ABA RESPOSTAS:",
            "   • ID: Número único do aluno",
            "   • Nome: Nome completo do aluno",
            "   • Sede: Unidade do colégio (opcional)",
            "   • Idioma escolhido: Inglês ou Espanhol",
            "   • Questão 01-70: Respostas do aluno (A, B, C, D ou E)",
            "",
            "2. ABA GABARITO:",
            "   • Questão: Número da questão (1-70)",
            "   • Disciplina: Nome da matéria",
            "   • Resposta: Resposta correta (A, B, C, D ou E)",
            "",
            "3. QUESTÕES DE LÍNGUAS:",
            "   • Questões 57-63: Inglês E Espanhol",
            "   • O sistema escolhe automaticamente baseado no 'Idioma escolhido'",
            "   • Ambas devem estar no gabarito",
            "",
            "4. FORMATO:",
            "   • Salve como .xlsx",
            "   • Não altere os nomes das abas",
            "   • Não altere os cabeçalhos",
            "",
            "5. SUPORTE:",
            "   • Em caso de dúvidas, consulte a documentação",
            "   • Sistema desenvolvido para Colégio Fleming",
            "",
            "Versão 2.0 - FastAPI + React"
        ]
        
        # Título
        cell = ws.cell(row=1, column=1, value=instrucoes[0])
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        # Instruções
        for row, instrucao in enumerate(instrucoes[1:], 2):
            cell = ws.cell(row=row, column=1, value=instrucao)
            if instrucao.startswith(('1.', '2.', '3.', '4.', '5.')):
                cell.font = Font(bold=True)
        
        # Ajustar largura
        ws.column_dimensions['A'].width = 80
    
    async def limpar_arquivos_temporarios(self, pdfs_info: List[PDFInfo]):
        """Limpa arquivos temporários"""
        
        try:
            # Remover PDFs individuais
            for pdf_info in pdfs_info:
                if os.path.exists(pdf_info.caminho):
                    os.remove(pdf_info.caminho)
            
            # Remover gráficos
            for arquivo in os.listdir(self.temp_dir):
                if arquivo.startswith('grafico_') and arquivo.endswith('.png'):
                    os.remove(os.path.join(self.temp_dir, arquivo))
            
        except Exception as e:
            print(f"Erro ao limpar arquivos temporários: {e}")
